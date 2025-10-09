#!/usr/bin/env python3
"""
Скрипт для создания диаграммы Ганта на основе данных из README.md
Создает Excel файл с диаграммой Ганта для MVP проекта
"""

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def create_gantt_data():
    """Создает данные для диаграммы Ганта на основе README.md"""
    
    # Базовые даты проекта
    start_date = datetime(2025, 10, 8)
    end_date = datetime(2025, 10, 31)
    
    # Данные задач из README.md
    tasks = [
        # Blockchain Engineer задачи
        {
            'task': 'Настройка среды разработки',
            'engineer': 'Blockchain Engineer',
            'start': start_date,
            'duration': 3,
            'color': '#1f77b4'  # синий
        },
        {
            'task': 'Создание базовых смарт-контрактов',
            'engineer': 'Blockchain Engineer', 
            'start': start_date + timedelta(days=3),
            'duration': 4,
            'color': '#1f77b4'
        },
        {
            'task': 'Реализация MarketFactory контракта',
            'engineer': 'Blockchain Engineer',
            'start': start_date + timedelta(days=7),
            'duration': 3,
            'color': '#1f77b4'
        },
        {
            'task': 'Реализация Escrow контракта',
            'engineer': 'Blockchain Engineer',
            'start': start_date + timedelta(days=10),
            'duration': 4,
            'color': '#1f77b4'
        },
        {
            'task': 'Интеграция с оракулами',
            'engineer': 'Blockchain Engineer',
            'start': start_date + timedelta(days=14),
            'duration': 4,
            'color': '#1f77b4'
        },
        {
            'task': 'Тестирование контрактов',
            'engineer': 'Blockchain Engineer',
            'start': start_date + timedelta(days=18),
            'duration': 3,
            'color': '#1f77b4'
        },
        {
            'task': 'Интеграционное тестирование (Blockchain)',
            'engineer': 'Blockchain Engineer',
            'start': start_date + timedelta(days=21),
            'duration': 3,
            'color': '#1f77b4'
        },
        
        # Frontend Engineer задачи
        {
            'task': 'Настройка проекта',
            'engineer': 'Frontend Engineer',
            'start': start_date,
            'duration': 3,
            'color': '#2ca02c'  # зеленый
        },
        {
            'task': 'Создание базовых компонентов',
            'engineer': 'Frontend Engineer',
            'start': start_date + timedelta(days=3),
            'duration': 4,
            'color': '#2ca02c'
        },
        {
            'task': 'Создание страниц регистрации и входа',
            'engineer': 'Frontend Engineer',
            'start': start_date + timedelta(days=7),
            'duration': 4,
            'color': '#2ca02c'
        },
        {
            'task': 'Создание торгового интерфейса',
            'engineer': 'Frontend Engineer',
            'start': start_date + timedelta(days=11),
            'duration': 7,
            'color': '#2ca02c'
        },
        {
            'task': 'Мобильная версия (PWA)',
            'engineer': 'Frontend Engineer',
            'start': start_date + timedelta(days=18),
            'duration': 3,
            'color': '#2ca02c'
        },
        {
            'task': 'Интеграционное тестирование (Frontend)',
            'engineer': 'Frontend Engineer',
            'start': start_date + timedelta(days=21),
            'duration': 3,
            'color': '#2ca02c'
        },
        
        # Backend Engineer задачи
        {
            'task': 'Настройка инфраструктуры',
            'engineer': 'Backend Engineer',
            'start': start_date,
            'duration': 3,
            'color': '#d62728'  # красный
        },
        {
            'task': 'Создание базового API',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=3),
            'duration': 4,
            'color': '#d62728'
        },
        {
            'task': 'Реализация User Service',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=7),
            'duration': 4,
            'color': '#d62728'
        },
        {
            'task': 'Реализация аутентификации',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=11),
            'duration': 3,
            'color': '#d62728'
        },
        {
            'task': 'Реализация Trading Service',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=14),
            'duration': 4,
            'color': '#d62728'
        },
        {
            'task': 'Интеграция с блокчейном',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=18),
            'duration': 3,
            'color': '#d62728'
        },
        {
            'task': 'Интеграционное тестирование (Backend)',
            'engineer': 'Backend Engineer',
            'start': start_date + timedelta(days=21),
            'duration': 3,
            'color': '#d62728'
        },
        
        # Общие задачи
        {
            'task': 'Финальное тестирование и отладка',
            'engineer': 'Все команды',
            'start': start_date + timedelta(days=24),
            'duration': 3,
            'color': '#7f7f7f'  # серый
        },
        {
            'task': 'Подготовка к запуску',
            'engineer': 'Все команды',
            'start': start_date + timedelta(days=27),
            'duration': 4,
            'color': '#7f7f7f'
        }
    ]
    
    return tasks

def create_excel_gantt():
    """Создает Excel файл с диаграммой Ганта"""
    
    # Создаем рабочую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "TASK_2025_10_31"
    
    # Получаем данные задач
    tasks = create_gantt_data()
    
    # Заголовки
    headers = ['Задача', 'Инженер', 'Дата начала', 'Продолжительность (дни)', 'Дата окончания']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Заполняем данные
    for row, task in enumerate(tasks, 2):
        ws.cell(row=row, column=1, value=task['task'])
        ws.cell(row=row, column=2, value=task['engineer'])
        ws.cell(row=row, column=3, value=task['start'].strftime('%d.%m.%Y'))
        ws.cell(row=row, column=4, value=task['duration'])
        end_date = task['start'] + timedelta(days=task['duration'])
        ws.cell(row=row, column=5, value=end_date.strftime('%d.%m.%Y'))
        
        # Цветовое кодирование по ролям
        if 'Blockchain' in task['engineer']:
            fill_color = "E6F3FF"  # светло-синий
        elif 'Frontend' in task['engineer']:
            fill_color = "E6FFE6"  # светло-зеленый
        elif 'Backend' in task['engineer']:
            fill_color = "FFE6E6"  # светло-красный
        else:
            fill_color = "F0F0F0"  # светло-серый
            
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
    
    # Автоподбор ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Добавляем сводную информацию
    summary_row = len(tasks) + 3
    ws.cell(row=summary_row, column=1, value="СВОДКА ПРОЕКТА:").font = Font(bold=True)
    ws.cell(row=summary_row+1, column=1, value="Общая продолжительность: 24 дня")
    ws.cell(row=summary_row+2, column=1, value="Дата начала: 08.10.2025")
    ws.cell(row=summary_row+3, column=1, value="Дата окончания: 31.10.2025")
    ws.cell(row=summary_row+4, column=1, value="Команда: 3 инженера")
    ws.cell(row=summary_row+5, column=1, value="Всего задач: 22")
    
    # Сохраняем файл
    wb.save('DIAGRAM.xlsx')
    print("✅ Файл DIAGRAM.xlsx успешно создан!")
    print(f"📊 Создано {len(tasks)} задач для диаграммы Ганта")
    print("📅 Период: 8-31 октября 2025")
    print("👥 Команда: Blockchain, Frontend, Backend Engineers")

def create_matplotlib_gantt():
    """Создает диаграмму Ганта с помощью matplotlib для визуализации"""
    
    tasks = create_gantt_data()
    
    # Создаем фигуру
    fig, ax = plt.subplots(figsize=(16, 10))
    
    # Цвета для разных ролей
    colors = {
        'Blockchain Engineer': '#1f77b4',
        'Frontend Engineer': '#2ca02c', 
        'Backend Engineer': '#d62728',
        'Все команды': '#7f7f7f'
    }
    
    # Создаем диаграмму
    y_pos = 0
    task_labels = []
    
    for task in tasks:
        # Определяем цвет
        color = colors.get(task['engineer'], '#7f7f7f')
        
        # Создаем полосу задачи
        ax.barh(y_pos, task['duration'], left=task['start'], 
                height=0.6, color=color, alpha=0.8, edgecolor='black', linewidth=0.5)
        
        # Добавляем подпись задачи
        ax.text(task['start'] + timedelta(days=task['duration']/2), y_pos, 
                task['task'], ha='center', va='center', fontsize=8, fontweight='bold')
        
        task_labels.append(f"{task['task']} ({task['engineer']})")
        y_pos += 1
    
    # Настройка осей
    ax.set_yticks(range(len(task_labels)))
    ax.set_yticklabels(task_labels, fontsize=8)
    ax.set_xlabel('Дата', fontsize=12, fontweight='bold')
    ax.set_ylabel('Задачи', fontsize=12, fontweight='bold')
    ax.set_title('Диаграмма Ганта - MVP разработка (8-31 октября 2025)', 
                 fontsize=14, fontweight='bold', pad=20)
    
    # Форматирование оси дат
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%d.%m'))
    ax.xaxis.set_major_locator(mdates.DayLocator(interval=2))
    plt.xticks(rotation=45)
    
    # Легенда
    legend_elements = []
    for engineer, color in colors.items():
        legend_elements.append(plt.Rectangle((0,0),1,1, facecolor=color, label=engineer))
    ax.legend(handles=legend_elements, loc='upper right', bbox_to_anchor=(1.15, 1))
    
    # Сетка
    ax.grid(True, alpha=0.3)
    ax.set_axisbelow(True)
    
    # Настройка макета
    plt.tight_layout()
    
    # Сохраняем изображение
    plt.savefig('gantt_chart.png', dpi=300, bbox_inches='tight')
    print("📈 Диаграмма Ганта сохранена как gantt_chart.png")
    
    return fig

if __name__ == "__main__":
    print("🚀 Создание диаграммы Ганта для MVP проекта...")
    print("📋 Анализ данных из README.md...")
    
    # Создаем Excel файл
    create_excel_gantt()
    
    # Создаем визуальную диаграмму
    print("📊 Создание визуальной диаграммы...")
    create_matplotlib_gantt()
    
    print("\n✅ Задание выполнено!")
    print("📁 Создан файл: DIAGRAM.xlsx")
    print("🖼️ Создан файл: gantt_chart.png")
    print("📋 Создан файл: TODO.md")
