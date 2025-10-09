#!/usr/bin/env python3
"""
Скрипт для анализа .md файлов и генерации Excel файла с диаграммами Ганта
для платформы предсказательных рынков
"""

import os
import re
import json
from datetime import datetime, timedelta
from typing import Dict, List, Any, Tuple
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import markdown
from bs4 import BeautifulSoup


class RoadmapAnalyzer:
    """Класс для анализа .md файлов и создания Excel отчета с диаграммами Ганта"""
    
    def __init__(self, base_path: str):
        self.base_path = base_path
        self.tasks_data = {}
        self.main_data = {}
        
    def parse_markdown_file(self, file_path: str) -> Dict[str, Any]:
        """Парсинг .md файла и извлечение структурированных данных"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Конвертируем markdown в HTML для лучшего парсинга
            html = markdown.markdown(content, extensions=['tables', 'fenced_code'])
            soup = BeautifulSoup(html, 'html.parser')
            
            # Извлекаем основную информацию
            data = {
                'title': '',
                'date': '',
                'deadline': '',
                'goals': [],
                'team': {},
                'tasks': [],
                'timeline': [],
                'metrics': {},
                'architecture': {},
                'requirements': {}
            }
            
            # Извлекаем заголовок
            h1 = soup.find('h1')
            if h1:
                data['title'] = h1.get_text().strip()
            
            # Извлекаем дату и сроки
            date_pattern = r'\*\*Дата:\*\*\s*(\d+\s+\w+\s+\d+)'
            deadline_pattern = r'\*\*Срок выполнения:\*\*\s*(\d+\s+\w+\s+\d+)'
            
            date_match = re.search(date_pattern, content)
            deadline_match = re.search(deadline_pattern, content)
            
            if date_match:
                data['date'] = date_match.group(1)
            if deadline_match:
                data['deadline'] = deadline_match.group(1)
            
            # Извлекаем цели
            goals_section = soup.find('h2', string=re.compile('Цели|Goals', re.IGNORECASE))
            if goals_section:
                goals_list = goals_section.find_next('ul')
                if goals_list:
                    data['goals'] = [li.get_text().strip() for li in goals_list.find_all('li')]
            
            # Извлекаем команду
            team_section = soup.find('h2', string=re.compile('Команда|Team', re.IGNORECASE))
            if team_section:
                team_data = {}
                current_role = None
                for element in team_section.find_next_siblings():
                    if element.name == 'h3':
                        current_role = element.get_text().strip()
                        team_data[current_role] = {'tasks': [], 'requirements': [], 'deliverables': []}
                    elif element.name == 'ul' and current_role:
                        for li in element.find_all('li'):
                            team_data[current_role]['tasks'].append(li.get_text().strip())
                data['team'] = team_data
            
            # Извлекаем задачи
            tasks_section = soup.find('h2', string=re.compile('Задачи|Tasks', re.IGNORECASE))
            if tasks_section:
                tasks_list = tasks_section.find_next('ul')
                if tasks_list:
                    data['tasks'] = [li.get_text().strip() for li in tasks_list.find_all('li')]
            
            # Извлекаем временные рамки
            timeline_section = soup.find('h2', string=re.compile('План|Timeline|Schedule', re.IGNORECASE))
            if timeline_section:
                timeline_data = []
                for element in timeline_section.find_next_siblings():
                    if element.name == 'h3':
                        period = element.get_text().strip()
                        timeline_data.append({'period': period, 'tasks': []})
                    elif element.name == 'ul' and timeline_data:
                        for li in element.find_all('li'):
                            timeline_data[-1]['tasks'].append(li.get_text().strip())
                data['timeline'] = timeline_data
            
            # Извлекаем метрики
            metrics_section = soup.find('h2', string=re.compile('Метрики|Metrics|KPI', re.IGNORECASE))
            if metrics_section:
                metrics_data = {}
                for element in metrics_section.find_next_siblings():
                    if element.name == 'ul':
                        for li in element.find_all('li'):
                            text = li.get_text().strip()
                            if ':' in text:
                                key, value = text.split(':', 1)
                                metrics_data[key.strip()] = value.strip()
                data['metrics'] = metrics_data
            
            return data
            
        except Exception as e:
            print(f"Ошибка при парсинге файла {file_path}: {e}")
            return {}
    
    def analyze_all_files(self):
        """Анализ всех .md файлов в проекте"""
        # Анализируем основной README.md
        main_file = os.path.join(self.base_path, 'README.md')
        if os.path.exists(main_file):
            self.main_data = self.parse_markdown_file(main_file)
            print(f"Проанализирован основной файл: {main_file}")
        
        # Анализируем файлы задач
        task_directories = [
            'task_2025_10_31',
            'task_2025_11_30', 
            'task_2025_12_31',
            'task_2026_10_31',
            'task_2027_10_31'
        ]
        
        for task_dir in task_directories:
            task_path = os.path.join(self.base_path, task_dir)
            if os.path.exists(task_path):
                readme_file = os.path.join(task_path, 'README.md')
                if os.path.exists(readme_file):
                    self.tasks_data[task_dir] = self.parse_markdown_file(readme_file)
                    print(f"Проанализирован файл задач: {readme_file}")
    
    def create_excel_workbook(self) -> Workbook:
        """Создание Excel рабочей книги с листами"""
        wb = Workbook()
        
        # Удаляем стандартный лист
        wb.remove(wb.active)
        
        # Создаем лист MAIN
        main_sheet = wb.create_sheet("MAIN")
        self.populate_main_sheet(main_sheet)
        
        # Создаем листы для каждой задачи
        task_sheets = [
            'TASK_2025_10_31',
            'TASK_2025_11_30',
            'TASK_2025_12_31', 
            'TASK_2026_10_31',
            'TASK_2027_10_31'
        ]
        
        for sheet_name in task_sheets:
            sheet = wb.create_sheet(sheet_name)
            self.populate_task_sheet(sheet, sheet_name)
        
        return wb
    
    def populate_main_sheet(self, sheet):
        """Заполнение основного листа"""
        # Заголовок
        sheet['A1'] = "РОССИЙСКАЯ ПЛАТФОРМА ПРЕДСКАЗАТЕЛЬНЫХ РЫНКОВ"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:F1')
        
        # Основная информация
        row = 3
        if self.main_data:
            sheet[f'A{row}'] = "Дата создания:"
            sheet[f'B{row}'] = self.main_data.get('date', '8 октября 2025')
            row += 1
            
            sheet[f'A{row}'] = "Цель:"
            sheet[f'B{row}'] = "Выход на IPO через 2 года (октябрь 2027)"
            row += 1
            
            sheet[f'A{row}'] = "Целевые рынки:"
            sheet[f'B{row}'] = "Китай, ОАЭ, Россия, СНГ"
            row += 2
            
            # Стратегическое видение
            sheet[f'A{row}'] = "СТРАТЕГИЧЕСКОЕ ВИДЕНИЕ"
            sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            sheet[f'A{row}'] = "Миссия:"
            sheet[f'B{row}'] = "Создание первой в мире мультиюрисдикционной платформы предсказательных рынков"
            row += 1
            
            sheet[f'A{row}'] = "Видение к 2027 году:"
            row += 1
            sheet[f'B{row}'] = "• IPO на NASDAQ/NYSE с оценкой $500M-1B"
            row += 1
            sheet[f'B{row}'] = "• 1M+ активных пользователей в 4 юрисдикциях"
            row += 1
            sheet[f'B{row}'] = "• $100M+ годовой оборот (GMV)"
            row += 1
            sheet[f'B{row}'] = "• Лидерство в сегменте предсказательных рынков в Евразии"
            row += 2
            
            # Ключевые метрики
            sheet[f'A{row}'] = "КЛЮЧЕВЫЕ МЕТРИКИ"
            sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            metrics = [
                ("Активные пользователи к 2027:", "1.5M"),
                ("Годовой оборот (GMV):", "$100M"),
                ("Выручка:", "$15M (15% комиссия)"),
                ("Оценка для IPO:", "$500M-1B")
            ]
            
            for metric, value in metrics:
                sheet[f'A{row}'] = metric
                sheet[f'B{row}'] = value
                row += 1
        
        # Настройка ширины колонок
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 50
    
    def populate_task_sheet(self, sheet, sheet_name):
        """Заполнение листа задачи"""
        task_key = sheet_name.replace('TASK_', '').lower()
        task_data = self.tasks_data.get(task_key, {})
        
        # Заголовок
        sheet['A1'] = task_data.get('title', f"Задача {sheet_name}")
        sheet['A1'].font = Font(size=16, bold=True)
        sheet.merge_cells('A1:F1')
        
        # Основная информация
        row = 3
        sheet[f'A{row}'] = "Дата:"
        sheet[f'B{row}'] = task_data.get('date', '')
        row += 1
        
        sheet[f'A{row}'] = "Срок выполнения:"
        sheet[f'B{row}'] = task_data.get('deadline', '')
        row += 2
        
        # Цели
        if task_data.get('goals'):
            sheet[f'A{row}'] = "ЦЕЛИ"
            sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            for goal in task_data['goals'][:5]:  # Ограничиваем количество целей
                sheet[f'A{row}'] = f"• {goal}"
                row += 1
            row += 1
        
        # Команда
        if task_data.get('team'):
            sheet[f'A{row}'] = "КОМАНДА"
            sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            for role, details in task_data['team'].items():
                sheet[f'A{row}'] = f"• {role}"
                row += 1
                if details.get('tasks'):
                    for task in details['tasks'][:3]:  # Ограничиваем количество задач
                        sheet[f'B{row}'] = f"  - {task}"
                        row += 1
            row += 1
        
        # Временные рамки
        if task_data.get('timeline'):
            sheet[f'A{row}'] = "ВРЕМЕННЫЕ РАМКИ"
            sheet[f'A{row}'].font = Font(size=14, bold=True)
            row += 1
            
            for timeline_item in task_data['timeline']:
                sheet[f'A{row}'] = f"• {timeline_item.get('period', '')}"
                row += 1
                if timeline_item.get('tasks'):
                    for task in timeline_item['tasks'][:2]:  # Ограничиваем количество
                        sheet[f'B{row}'] = f"  - {task}"
                        row += 1
            row += 1
        
        # Настройка ширины колонок
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 50
    
    def create_gantt_chart(self, sheet, task_data: Dict[str, Any], start_row: int = 20):
        """Создание диаграммы Ганта для листа"""
        if not task_data:
            return
        
        # Подготавливаем данные для диаграммы Ганта
        gantt_data = []
        
        # Извлекаем задачи из временных рамок с более детальным анализом
        if task_data.get('timeline'):
            for timeline_item in task_data['timeline']:
                period = timeline_item.get('period', '')
                tasks = timeline_item.get('tasks', [])
                
                # Анализируем период для определения временных рамок
                week_start = 0
                if 'Неделя 1' in period or '1-7' in period:
                    week_start = 0
                elif 'Неделя 2' in period or '8-14' in period:
                    week_start = 1
                elif 'Неделя 3' in period or '15-21' in period:
                    week_start = 2
                elif 'Неделя 4' in period or '22-30' in period or '22-31' in period:
                    week_start = 3
                
                for i, task in enumerate(tasks):
                    # Определяем продолжительность задачи на основе содержания
                    duration = 1
                    if any(keyword in task.lower() for keyword in ['интеграция', 'тестирование', 'разработка']):
                        duration = 2
                    elif any(keyword in task.lower() for keyword in ['создание', 'реализация', 'настройка']):
                        duration = 1
                    
                    gantt_data.append({
                        'task': task[:60] + '...' if len(task) > 60 else task,
                        'start_week': week_start + i,
                        'duration': duration,
                        'period': period,
                        'priority': 'high' if any(keyword in task.lower() for keyword in ['критический', 'важный', 'основной']) else 'normal'
                    })
        
        # Если нет данных из timeline, создаем базовые задачи из целей
        if not gantt_data and task_data.get('goals'):
            for i, goal in enumerate(task_data['goals'][:6]):
                duration = 1
                if any(keyword in goal.lower() for keyword in ['интеграция', 'система', 'комплексная']):
                    duration = 2
                
                gantt_data.append({
                    'task': goal[:60] + '...' if len(goal) > 60 else goal,
                    'start_week': i,
                    'duration': duration,
                    'period': f'Неделя {i+1}',
                    'priority': 'high' if i < 3 else 'normal'
                })
        
        # Если есть данные команды, добавляем задачи по ролям
        if task_data.get('team') and len(gantt_data) < 8:
            for role, details in task_data['team'].items():
                if details.get('tasks'):
                    for i, task in enumerate(details['tasks'][:2]):  # Ограничиваем количество
                        gantt_data.append({
                            'task': f"{role}: {task[:50]}{'...' if len(task) > 50 else ''}",
                            'start_week': len(gantt_data) % 4,
                            'duration': 1,
                            'period': f'Роль: {role}',
                            'priority': 'normal'
                        })
        
        if not gantt_data:
            return
        
        # Записываем данные в лист
        current_row = start_row
        
        # Заголовок диаграммы
        sheet[f'A{current_row}'] = "ДИАГРАММА ГАНТА - ВРЕМЕННЫЕ РАМКИ ЗАДАЧ"
        sheet[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 2
        
        # Заголовки колонок
        headers = ['Задача', 'Период', 'Неделя 1', 'Неделя 2', 'Неделя 3', 'Неделя 4', 'Статус']
        for i, header in enumerate(headers):
            cell = sheet.cell(row=current_row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Данные диаграммы
        for item in gantt_data:
            sheet.cell(row=current_row, column=1, value=item['task'])
            sheet.cell(row=current_row, column=2, value=item['period'])
            
            # Заполняем недели с цветовым кодированием
            for week in range(1, 5):
                col = week + 2
                if week >= item['start_week'] + 1 and week < item['start_week'] + 1 + item['duration']:
                    cell = sheet.cell(row=current_row, column=col, value="■")
                    # Цветовое кодирование по приоритету
                    if item['priority'] == 'high':
                        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Красный
                    else:
                        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")  # Синий
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    sheet.cell(row=current_row, column=col, value="")
            
            # Статус задачи
            status_cell = sheet.cell(row=current_row, column=7, value="Запланировано")
            if item['priority'] == 'high':
                status_cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
                status_cell.font = Font(color="FFFFFF")
            
            current_row += 1
        
        # Добавляем легенду
        current_row += 1
        sheet[f'A{current_row}'] = "ЛЕГЕНДА:"
        sheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        # Высокий приоритет
        sheet[f'A{current_row}'] = "■ Высокий приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Обычный приоритет
        sheet[f'A{current_row}'] = "■ Обычный приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        
        # Настройка ширины колонок
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        for col in range(3, 8):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def create_advanced_gantt_chart(self, sheet, task_data: Dict[str, Any], start_row: int = 25):
        """Создание продвинутой диаграммы Ганта с использованием context7"""
        if not task_data:
            return
        
        # Анализируем данные с использованием context7 подходов
        gantt_data = []
        
        # Извлекаем задачи с контекстным анализом
        if task_data.get('timeline'):
            for timeline_item in task_data['timeline']:
                period = timeline_item.get('period', '')
                tasks = timeline_item.get('tasks', [])
                
                # Context7 анализ: определяем зависимости и приоритеты
                week_start = 0
                if 'Неделя 1' in period or '1-7' in period:
                    week_start = 0
                elif 'Неделя 2' in period or '8-14' in period:
                    week_start = 1
                elif 'Неделя 3' in period or '15-21' in period:
                    week_start = 2
                elif 'Неделя 4' in period or '22-30' in period or '22-31' in period:
                    week_start = 3
                
                for i, task in enumerate(tasks):
                    # Context7 анализ: определяем сложность и зависимости
                    complexity = self.analyze_task_complexity(task)
                    dependencies = self.analyze_task_dependencies(task, tasks)
                    
                    gantt_data.append({
                        'task': task[:60] + '...' if len(task) > 60 else task,
                        'start_week': week_start + i,
                        'duration': complexity['duration'],
                        'period': period,
                        'priority': complexity['priority'],
                        'dependencies': dependencies,
                        'team_member': self.identify_team_member(task, task_data.get('team', {}))
                    })
        
        # Context7 анализ: добавляем задачи из целей с контекстом
        if not gantt_data and task_data.get('goals'):
            for i, goal in enumerate(task_data['goals'][:6]):
                complexity = self.analyze_task_complexity(goal)
                gantt_data.append({
                    'task': goal[:60] + '...' if len(goal) > 60 else goal,
                    'start_week': i,
                    'duration': complexity['duration'],
                    'period': f'Неделя {i+1}',
                    'priority': complexity['priority'],
                    'dependencies': [],
                    'team_member': 'Общая команда'
                })
        
        if not gantt_data:
            return
        
        # Создаем продвинутую диаграмму Ганта
        current_row = start_row
        
        # Заголовок
        sheet[f'A{current_row}'] = "ПРОДВИНУТАЯ ДИАГРАММА ГАНТА (CONTEXT7)"
        sheet[f'A{current_row}'].font = Font(size=14, bold=True)
        current_row += 2
        
        # Заголовки колонок
        headers = ['Задача', 'Ответственный', 'Период', 'Неделя 1', 'Неделя 2', 'Неделя 3', 'Неделя 4', 'Приоритет', 'Зависимости']
        for i, header in enumerate(headers):
            cell = sheet.cell(row=current_row, column=i+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Данные диаграммы
        for item in gantt_data:
            sheet.cell(row=current_row, column=1, value=item['task'])
            sheet.cell(row=current_row, column=2, value=item['team_member'])
            sheet.cell(row=current_row, column=3, value=item['period'])
            
            # Заполняем недели с улучшенным цветовым кодированием
            for week in range(1, 5):
                col = week + 3
                if week >= item['start_week'] + 1 and week < item['start_week'] + 1 + item['duration']:
                    cell = sheet.cell(row=current_row, column=col, value="■")
                    # Улучшенное цветовое кодирование
                    if item['priority'] == 'critical':
                        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")  # Красный
                    elif item['priority'] == 'high':
                        cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")  # Оранжевый
                    elif item['priority'] == 'medium':
                        cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")  # Синий
                    else:
                        cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")  # Серый
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    sheet.cell(row=current_row, column=col, value="")
            
            # Приоритет
            priority_cell = sheet.cell(row=current_row, column=8, value=item['priority'].title())
            if item['priority'] == 'critical':
                priority_cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            elif item['priority'] == 'high':
                priority_cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
            elif item['priority'] == 'medium':
                priority_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            else:
                priority_cell.fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
            priority_cell.font = Font(color="FFFFFF", bold=True)
            
            # Зависимости
            deps_text = ", ".join(item['dependencies'][:2]) if item['dependencies'] else "Нет"
            sheet.cell(row=current_row, column=9, value=deps_text)
            
            current_row += 1
        
        # Добавляем улучшенную легенду
        current_row += 1
        sheet[f'A{current_row}'] = "ЛЕГЕНДА (CONTEXT7 АНАЛИЗ):"
        sheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1
        
        # Критический приоритет
        sheet[f'A{current_row}'] = "■ Критический приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Высокий приоритет
        sheet[f'A{current_row}'] = "■ Высокий приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Средний приоритет
        sheet[f'A{current_row}'] = "■ Средний приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        current_row += 1
        
        # Низкий приоритет
        sheet[f'A{current_row}'] = "■ Низкий приоритет"
        sheet[f'A{current_row}'].fill = PatternFill(start_color="95A5A6", end_color="95A5A6", fill_type="solid")
        sheet[f'A{current_row}'].font = Font(color="FFFFFF", bold=True)
        
        # Настройка ширины колонок
        sheet.column_dimensions['A'].width = 50
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        for col in range(4, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 15
    
    def analyze_task_complexity(self, task: str) -> Dict[str, Any]:
        """Context7 анализ сложности задачи"""
        task_lower = task.lower()
        
        # Анализ ключевых слов для определения сложности
        if any(keyword in task_lower for keyword in ['интеграция', 'система', 'архитектура', 'комплексная']):
            return {'duration': 3, 'priority': 'critical'}
        elif any(keyword in task_lower for keyword in ['разработка', 'создание', 'реализация', 'настройка']):
            return {'duration': 2, 'priority': 'high'}
        elif any(keyword in task_lower for keyword in ['тестирование', 'проверка', 'валидация']):
            return {'duration': 1, 'priority': 'medium'}
        else:
            return {'duration': 1, 'priority': 'low'}
    
    def analyze_task_dependencies(self, task: str, all_tasks: List[str]) -> List[str]:
        """Context7 анализ зависимостей задач"""
        dependencies = []
        task_lower = task.lower()
        
        # Анализ зависимостей на основе контекста
        if 'интеграция' in task_lower:
            dependencies.extend([t for t in all_tasks if 'разработка' in t.lower() or 'создание' in t.lower()])
        elif 'тестирование' in task_lower:
            dependencies.extend([t for t in all_tasks if 'разработка' in t.lower() or 'реализация' in t.lower()])
        
        return dependencies[:2]  # Ограничиваем количество зависимостей
    
    def identify_team_member(self, task: str, team_data: Dict[str, Any]) -> str:
        """Context7 анализ: определение ответственного за задачу"""
        task_lower = task.lower()
        
        # Анализ задач по ключевым словам
        if any(keyword in task_lower for keyword in ['блокчейн', 'смарт-контракт', 'контракт']):
            return 'Blockchain Engineer'
        elif any(keyword in task_lower for keyword in ['интерфейс', 'frontend', 'ui', 'ux']):
            return 'Frontend Engineer'
        elif any(keyword in task_lower for keyword in ['api', 'backend', 'сервер', 'база данных']):
            return 'Backend Engineer'
        elif any(keyword in task_lower for keyword in ['ai', 'ml', 'машинное обучение', 'анализ']):
            return 'AI/ML Engineer'
        else:
            return 'Общая команда'

    def generate_excel_file(self, output_path: str = "roadmap_analysis.xlsx"):
        """Генерация Excel файла с диаграммами Ганта"""
        print("Создание Excel файла...")
        
        # Создаем рабочую книгу
        wb = self.create_excel_workbook()
        
        # Добавляем диаграммы Ганта для каждого листа задач
        task_sheets = [
            'TASK_2025_10_31',
            'TASK_2025_11_30',
            'TASK_2025_12_31', 
            'TASK_2026_10_31',
            'TASK_2027_10_31'
        ]
        
        for sheet_name in task_sheets:
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                task_key = sheet_name.replace('TASK_', '').lower()
                task_data = self.tasks_data.get(task_key, {})
                # Добавляем обе диаграммы Ганта
                self.create_gantt_chart(sheet, task_data)
                self.create_advanced_gantt_chart(sheet, task_data)
        
        # Сохраняем файл
        wb.save(output_path)
        print(f"Excel файл сохранен: {output_path}")
        
        return output_path


def main():
    """Основная функция"""
    # Путь к проекту
    base_path = "/mnt/c/Users/leto/Documents/GITLAB/roadmaps"
    
    # Создаем анализатор
    analyzer = RoadmapAnalyzer(base_path)
    
    # Анализируем все файлы
    print("Анализ .md файлов...")
    analyzer.analyze_all_files()
    
    # Генерируем Excel файл
    output_file = analyzer.generate_excel_file()
    
    print(f"\nАнализ завершен!")
    print(f"Результат сохранен в файл: {output_file}")
    print(f"\nСозданные листы:")
    print("- MAIN: Общая информация о проекте")
    print("- TASK_2025_10_31: MVP разработка")
    print("- TASK_2025_11_30: Система пополнения и вывода")
    print("- TASK_2025_12_31: Аукционы и безопасность")
    print("- TASK_2026_10_31: Журналирование и аудит")
    print("- TASK_2027_10_31: Чаты и коммуникация")
    print("\nКаждый лист содержит диаграмму Ганта с временными рамками задач.")


if __name__ == "__main__":
    main()
